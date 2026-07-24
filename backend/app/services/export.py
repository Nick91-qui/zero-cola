from io import BytesIO
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


class ExportService:
    @staticmethod
    def generate_exam_pdf_report(
        exam_title: str,
        class_id: str,
        teacher_name: str,
        max_score: float,
        total_questions: int,
        attempts: List[Dict[str, Any]],
        question_stats: List[Dict[str, Any]],
    ) -> bytes:
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36,
        )
        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "TitleStyle",
            parent=styles["Heading1"],
            fontSize=18,
            leading=22,
            textColor=colors.HexColor("#1e293b"),
            alignment=1,
            spaceAfter=12,
        )
        subtitle_style = ParagraphStyle(
            "SubTitleStyle",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#475569"),
            alignment=1,
            spaceAfter=16,
        )
        section_style = ParagraphStyle(
            "SectionStyle",
            parent=styles["Heading2"],
            fontSize=13,
            leading=16,
            textColor=colors.HexColor("#0f172a"),
            spaceBefore=14,
            spaceAfter=8,
        )

        story.append(Paragraph(f"Relatório de Avaliação: {exam_title}", title_style))
        story.append(
            Paragraph(
                f"Turma: {class_id or 'Geral'} | Professor: {teacher_name} | Questões: {total_questions} | Nota Máxima: {max_score}",
                subtitle_style,
            )
        )

        # 1. General Summary
        total_students = len(attempts)
        avg_score = (
            sum(a["final_score"] for a in attempts) / total_students
            if total_students > 0
            else 0.0
        )
        avg_accuracy = (
            sum(a["accuracy_percentage"] for a in attempts) / total_students
            if total_students > 0
            else 0.0
        )

        summary_data = [
            ["Total de Alunos", "Média da Turma", "Precisão Média da Turma"],
            [f"{total_students}", f"{avg_score:.2f} / {max_score:.2f}", f"{avg_accuracy:.1f}%"],
        ]
        summary_table = Table(summary_data, colWidths=[170, 170, 170])
        summary_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e2e8f0")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                    ("PADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(summary_table)
        story.append(Spacer(1, 14))

        # 2. Student Results Table
        story.append(Paragraph("Desempenho por Aluno", section_style))
        student_table_data = [["Código / ID", "Aluno", "Acertos", "Precisão (%)", "Nota Final"]]
        for att in attempts:
            student_table_data.append(
                [
                    att.get("student_code") or "-",
                    att.get("student_name") or "Aluno",
                    f"{att['correct_answers']} / {total_questions}",
                    f"{att['accuracy_percentage']:.1f}%",
                    f"{att['final_score']:.2f}",
                ]
            )

        student_table = Table(student_table_data, colWidths=[90, 180, 80, 80, 80])
        student_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                    ("PADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(student_table)
        story.append(Spacer(1, 14))

        # 3. Question Statistics Table
        story.append(Paragraph("Estatísticas por Questão", section_style))
        q_table_data = [["Questão", "Gabarito", "Respostas", "Acertos", "Erros", "% Acerto"]]
        for qs in question_stats:
            q_table_data.append(
                [
                    f"Q{qs['question_number']}",
                    qs.get("correct_option") or "-",
                    f"{qs['total_responses']}",
                    f"{qs['correct_count']}",
                    f"{qs['incorrect_count']}",
                    f"{qs['accuracy_percentage']:.1f}%",
                ]
            )

        q_table = Table(q_table_data, colWidths=[65, 65, 80, 75, 75, 150])
        q_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#475569")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                    ("PADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        story.append(q_table)

        doc.build(story)
        pdf_data = buffer.getvalue()
        buffer.close()
        return pdf_data

    @staticmethod
    def generate_exam_xlsx_report(
        exam_title: str,
        class_id: str,
        teacher_name: str,
        max_score: float,
        total_questions: int,
        attempts: List[Dict[str, Any]],
        question_stats: List[Dict[str, Any]],
    ) -> bytes:
        wb = openpyxl.Workbook()
        
        # Sheet 1: Alunos e Notas
        ws1 = wb.active
        ws1.title = "Resultados dos Alunos"

        # Headers
        ws1.append(["COLA-ZERO — Relatório da Avaliação"])
        ws1.append(["Título da Avaliação:", exam_title])
        ws1.append(["Turma:", class_id or "Geral"])
        ws1.append(["Professor:", teacher_name])
        ws1.append(["Total de Questões:", total_questions, "Nota Máxima:", max_score])
        ws1.append([])

        header_row = ["Código do Aluno", "Nome do Aluno", "Acertos", "Erros", "Precisão (%)", "Nota Final"]
        ws1.append(header_row)

        for att in attempts:
            ws1.append(
                [
                    att.get("student_code") or "-",
                    att.get("student_name") or "Aluno",
                    att["correct_answers"],
                    att["incorrect_answers"],
                    round(att["accuracy_percentage"], 1),
                    round(att["final_score"], 2),
                ]
            )

        # Sheet 2: Análise por Questão
        ws2 = wb.create_sheet(title="Estatísticas por Questão")
        ws2.append(["Análise Pedagógica por Questão — " + exam_title])
        ws2.append([])
        ws2.append(["Questão", "Alternativa Correta", "Total Respostas", "Nº Acertos", "Nº Erros", "Precisão (%)", "Taxa Erro (%)"])

        for qs in question_stats:
            ws2.append(
                [
                    f"Q{qs['question_number']}",
                    qs.get("correct_option") or "-",
                    qs["total_responses"],
                    qs["correct_count"],
                    qs["incorrect_count"],
                    round(qs["accuracy_percentage"], 1),
                    round(qs["error_percentage"], 1),
                ]
            )

        output = BytesIO()
        wb.save(output)
        xlsx_bytes = output.getvalue()
        output.close()
        return xlsx_bytes

    @staticmethod
    def generate_student_pdf_report(
        student_name: str,
        student_code: str,
        exam_title: str,
        attempt: Dict[str, Any],
        answers: List[Dict[str, Any]],
    ) -> bytes:
        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36,
        )
        story = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "StudentTitle",
            parent=styles["Heading1"],
            fontSize=16,
            leading=20,
            textColor=colors.HexColor("#1e293b"),
            spaceAfter=8,
        )
        sub_style = ParagraphStyle(
            "StudentSub",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#475569"),
            spaceAfter=14,
        )

        story.append(Paragraph(f"Boletim Individual — {student_name}", title_style))
        story.append(
            Paragraph(
                f"Código: {student_code or '-'} | Avaliação: {exam_title} | Data: {attempt.get('completed_at', '-')}",
                sub_style,
            )
        )

        # Summary box
        summary_data = [
            ["Nota Final", "Acertos", "Precisão"],
            [
                f"{attempt['final_score']:.2f}",
                f"{attempt['correct_answers']} / {attempt['total_questions']}",
                f"{attempt['accuracy_percentage']:.1f}%",
            ],
        ]
        sum_table = Table(summary_data, colWidths=[170, 170, 170])
        sum_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                    ("PADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        story.append(sum_table)
        story.append(Spacer(1, 14))

        # Itemized answers
        ans_table_data = [["Questão", "Opção Selecionada", "Opção Correta", "Resultado"]]
        for ans in answers:
            res_str = "CORRETO" if ans.get("is_correct") else "ERRADO"
            ans_table_data.append(
                [
                    f"Q{ans['question_number']}",
                    ans.get("selected_option") or "-",
                    ans.get("correct_option") or "-",
                    res_str,
                ]
            )

        ans_table = Table(ans_table_data, colWidths=[100, 140, 140, 130])
        ans_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#475569")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                    ("PADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(ans_table)

        doc.build(story)
        pdf_data = buffer.getvalue()
        buffer.close()
        return pdf_data
